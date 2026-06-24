"""
RecorridosIA — Generador de informe FOR FO 02
Variables de entorno requeridas en Render:
    BOT_TOKEN       = token del bot de Telegram (@RecorridosIA_bot)
    GEMINI_API_KEY  = clave de API de Google Gemini AI
    MAPILLARY_TOKEN = token de acceso a Mapillary
    TOTP_SECRET     = clave secreta para proteger entrega del Excel (2FA)
"""

import os
import io
import hmac
import struct
import time
import base64
import hashlib
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

# ─── Variables de entorno (se configuran en Render) ───────────────────────────
BOT_TOKEN       = os.getenv("BOT_TOKEN")
GEMINI_API_KEY  = os.getenv("GEMINI_API_KEY")
MAPILLARY_TOKEN = os.getenv("MAPILLARY_TOKEN")
TOTP_SECRET     = os.getenv("TOTP_SECRET")


# ══════════════════════════════════════════════════════════════════════════════
#  TOTP — VERIFICACIÓN 2FA PARA ENTREGA DEL EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def generar_codigo_totp() -> str:
    """Genera el código TOTP actual de 6 dígitos (válido 30 segundos)."""
    secreto = base64.b32decode(TOTP_SECRET.upper().replace(" ", ""), casefold=True)
    contador = struct.pack(">Q", int(time.time()) // 30)
    mac = hmac.new(secreto, contador, hashlib.sha1).digest()
    offset = mac[-1] & 0x0F
    codigo = struct.unpack(">I", mac[offset:offset + 4])[0] & 0x7FFFFFFF
    return str(codigo % 1_000_000).zfill(6)


def verificar_codigo_totp(codigo_ingresado: str) -> bool:
    """
    Verifica el código ingresado por el técnico.
    Acepta el código actual y el anterior (ventana de 60 segundos).
    """
    secreto = base64.b32decode(TOTP_SECRET.upper().replace(" ", ""), casefold=True)
    ahora   = int(time.time()) // 30
    for delta in [0, -1, 1]:
        contador = struct.pack(">Q", ahora + delta)
        mac      = hmac.new(secreto, contador, hashlib.sha1).digest()
        offset   = mac[-1] & 0x0F
        codigo   = struct.unpack(">I", mac[offset:offset + 4])[0] & 0x7FFFFFFF
        if str(codigo % 1_000_000).zfill(6) == str(codigo_ingresado).strip():
            return True
    return False

# ─── Plantilla base incluida en el repositorio ────────────────────────────────
PLANTILLA_PATH = os.path.join(os.path.dirname(__file__), "plantilla_FOR_FO_02.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
#  ESTRUCTURA DE DATOS DEL INFORME
# ══════════════════════════════════════════════════════════════════════════════

def datos_vacios():
    """Retorna la estructura vacía que el bot irá llenando campo a campo."""
    return {
        # ── Pestaña: REPORTES_DE_RECORRIDOS ──────────────────────────────────
        "recorrido": {
            "fecha":           "",          # 22/05/2026
            "hora_inicio":     "",          # 15:40:21
            "hora_fin":        "",          # 16:45:19
            "nombre_ruta":     "",          # GOSSEAL-MACHACHI   TAREA: 157415066
            "codigo_cuadrilla":"",          # FO UIO INT 04
            "nodo_inicial":    "",          # GOSSEAL
            "nodo_final":      "",          # MACHACHI
            "lider":           "",          # RICHARD DAVID TAIPE COYAGO
            "ayudante":        "",          # JOSE LUIS ALLAICA CONDO
            "coordinador":     "",          # JUAN CARLOS YEPEZ ACAN
            "fotos_total":     0,
            "observaciones":   "",
            "novedades":       [],          # lista de dicts (ver novedad_vacia)
        },
        # ── Pestaña: Checklist CIU ────────────────────────────────────────────
        "ciu": {
            "distancia_ruta":  "",          # 59KM
            "vehiculo_placa":  "",          # PCO3940
            "herramientas":    {},          # nombre → {"cantidad": N, "obs": ""}
        },
        # ── Pestaña: Checklists MPRIU ─────────────────────────────────────────
        "mpriu": {
            "novedades_check": {},          # nombre_novedad → {"check": SI/NO, "cantidad": N}
            "observaciones":   "",
        },
        # ── Pestaña: MANGAS (solo cuando hay cambio de manga) ─────────────────
        "mangas": [],                       # lista de dicts (ver manga_vacia)
        # ── Pestaña: INVENTARIO DE HILOS EN NODO (solo cuando hay cambio ODF) ─
        "hilos": {
            "posicion_odf": "",
            "filas":        [],             # lista de dicts por fila
        },
    }


def novedad_vacia(numero: int) -> dict:
    return {
        "numero":      numero,
        "fecha":       "",
        "hora_inicio": "",
        "hora_fin":    "",
        "motivo":      "",
        "remedio":     "",
        "tarea_pendiente": "",
        "coordenadas": "",
        "foto_antes":  None,   # bytes de la imagen
        "foto_despues":None,
    }


def manga_vacia() -> dict:
    return {
        "nombre":      "",
        "derivacion":  "NO",
        "coordenadas": "",
        "observacion": "",
        "foto":        None,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  GENERADOR DEL EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def generar_excel(datos: dict) -> bytes:
    """
    Recibe el dict de datos completos y devuelve el .xlsx como bytes
    para enviarlo directo por Telegram sin guardar en disco.
    """
    wb = load_workbook(PLANTILLA_PATH)
    _llenar_recorridos(wb, datos)
    _llenar_fotos(wb, datos)
    _llenar_ciu(wb, datos)
    _llenar_mpriu(wb, datos)
    if datos["mangas"]:
        _llenar_mangas(wb, datos)
    if datos["hilos"]["filas"]:
        _llenar_hilos(wb, datos)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ─── Pestaña REPORTES_DE_RECORRIDOS ──────────────────────────────────────────

def _llenar_recorridos(wb, datos):
    ws = wb["REPORTES_DE_RECORRIDOS"]
    r  = datos["recorrido"]

    # Encabezado principal
    _set(ws, "C5",  r["fecha"])
    _set(ws, "D5",  r["hora_inicio"])
    _set(ws, "E5",  r["hora_fin"])
    _set(ws, "C6",  r["nombre_ruta"])
    _set(ws, "C7",  r["codigo_cuadrilla"])
    _set(ws, "C8",  r["nodo_inicial"])

    # Novedades (filas dinámicas, 4 filas por novedad a partir de fila 10)
    fila_base = 10
    for i, nov in enumerate(r["novedades"][:20]):
        offset = i * 7          # cada bloque ocupa ~7 filas en la plantilla
        f = fila_base + offset
        _set(ws, f"C{f}",   nov["fecha"])
        _set(ws, f"D{f}",   nov["hora_inicio"])
        _set(ws, f"E{f}",   nov["hora_fin"])
        _set(ws, f"C{f+1}", nov["motivo"])
        _set(ws, f"C{f+2}", nov["remedio"])
        _set(ws, f"C{f+3}", nov["tarea_pendiente"])
        _set(ws, f"C{f+4}", nov["coordenadas"])

    # Pie del reporte
    _set(ws, "C152", r["nodo_final"])
    _set(ws, "C153", r["lider"])
    _set(ws, "C154", r["ayudante"])
    _set(ws, "C155", r["coordinador"])
    _set(ws, "C156", str(r["fotos_total"]))
    _set(ws, "C157", r["observaciones"])


# ─── Pestaña FOTOS_ANEXAS_AL_REPORTE ─────────────────────────────────────────

def _llenar_fotos(wb, datos):
    ws  = wb["FOTOS_ANEXAS_AL_REPORTE"]
    r   = datos["recorrido"]
    from openpyxl.drawing.image import Image as XLImage

    fila_base = 10
    for i, nov in enumerate(r["novedades"][:20]):
        offset = i * 5
        f = fila_base + offset
        for col, foto in [("C", nov["foto_antes"]), ("E", nov["foto_despues"])]:
            if foto:
                try:
                    img = XLImage(io.BytesIO(foto))
                    img.width, img.height = 160, 120
                    ws.add_image(img, f"{col}{f}")
                except Exception:
                    pass


# ─── Pestaña Checklist CIU ────────────────────────────────────────────────────

def _llenar_ciu(wb, datos):
    ws  = wb["Checklist CIU"]
    r   = datos["recorrido"]
    ciu = datos["ciu"]

    _set(ws, "C3", r["fecha"])
    _set(ws, "F3", r["hora_inicio"])
    _set(ws, "H3", r["hora_fin"])
    _set(ws, "C4", r["nombre_ruta"])
    _set(ws, "C5", r["nodo_inicial"])
    _set(ws, "C6", r["nodo_final"])
    _set(ws, "C7", ciu.get("distancia_ruta", ""))
    _set(ws, "C8", r["lider"])
    _set(ws, "C9", ciu.get("vehiculo_placa", ""))
    _set(ws, "C10", r["coordinador"])

    # Herramientas y EPP (filas 13 en adelante)
    fila = 13
    for nombre, vals in ciu.get("herramientas", {}).items():
        for row in ws.iter_rows(min_row=fila, max_row=fila + 40):
            celda_nombre = row[1].value
            if celda_nombre and nombre.upper() in str(celda_nombre).upper():
                row[2].value = vals.get("cantidad", 0)
                row[3].value = vals.get("obs", "")
                break


# ─── Pestaña Checklists MPRIU ─────────────────────────────────────────────────

def _llenar_mpriu(wb, datos):
    ws    = wb["Checklists MPRIU"]
    r     = datos["recorrido"]
    mpriu = datos["mpriu"]

    _set(ws, "C3",  r["fecha"])
    _set(ws, "F3",  r["hora_inicio"])
    _set(ws, "H3",  r["hora_fin"])
    _set(ws, "C4",  r["nombre_ruta"])
    _set(ws, "C5",  r["nodo_inicial"])
    _set(ws, "C6",  r["nodo_final"])
    _set(ws, "C7",  datos["ciu"].get("distancia_ruta", ""))
    _set(ws, "C8",  r["lider"])
    _set(ws, "C9",  datos["ciu"].get("vehiculo_placa", ""))
    _set(ws, "C10", r["coordinador"])

    # Marcar SI/NO y cantidad en cada novedad del checklist
    for row in ws.iter_rows(min_row=13, max_row=60):
        celda_novedad = row[1].value
        if celda_novedad:
            key = str(celda_novedad).strip().upper()
            if key in mpriu.get("novedades_check", {}):
                val = mpriu["novedades_check"][key]
                row[2].value = "SI" if val.get("check") else "NO"
                row[6].value = val.get("cantidad", 0)

    _set(ws, "C62", mpriu.get("observaciones", r["observaciones"]))


# ─── Pestaña MANGAS ───────────────────────────────────────────────────────────

def _llenar_mangas(wb, datos):
    ws = wb["MANGAS"]
    mangas = datos["mangas"]
    fila = 7
    for i, m in enumerate(mangas):
        col_offset = 0 if i % 2 == 0 else 5
        f = fila + (i // 2) * 6
        ws.cell(row=f,   column=2 + col_offset).value = m["nombre"]
        ws.cell(row=f+1, column=2 + col_offset).value = m["derivacion"]
        ws.cell(row=f+2, column=2 + col_offset).value = m["coordenadas"]
        ws.cell(row=f+3, column=2 + col_offset).value = m["observacion"]


# ─── Pestaña INVENTARIO DE HILOS EN NODO ─────────────────────────────────────

def _llenar_hilos(wb, datos):
    ws    = wb["INVENTARIO DE HILOS EN NODO"]
    hilos = datos["hilos"]
    _set(ws, "G12", hilos.get("posicion_odf", ""))
    for i, fila_data in enumerate(hilos["filas"]):
        f = 5 + i
        ws.cell(row=f, column=2).value = fila_data.get("hilo_par",    "")
        ws.cell(row=f, column=3).value = fila_data.get("hilo_impar",  "")
        ws.cell(row=f, column=4).value = fila_data.get("descripcion", "")
        ws.cell(row=f, column=5).value = fila_data.get("estado",      "")


# ─── Helper ───────────────────────────────────────────────────────────────────

def _set(ws, celda: str, valor):
    """Escribe un valor en la celda preservando el formato existente."""
    ws[celda] = valor


# ══════════════════════════════════════════════════════════════════════════════
#  FLUJO DEL BOT — PREGUNTAS POR PESTAÑA
# ══════════════════════════════════════════════════════════════════════════════

"""
El bot (bot/main.py) sigue este orden de preguntas al técnico:

PESTAÑAS QUE SE LLENAN CADA RECORRIDO (siempre):
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  /inspeccionar  →  el bot pregunta:
  1. ¿Nombre de la ruta?
  2. ¿Código de cuadrilla?
  3. ¿Nodo inicial y final?
  4. ¿Líder, ayudante, coordinador?
  5. ¿Placa del vehículo?
  6. ¿Distancia de la ruta (km)?
  [La IA detecta novedades automáticamente del video]
  7. Para cada novedad: ¿tarea pendiente? + foto antes/después
  8. ¿Observaciones generales?
  → Genera Excel y lo envía por Telegram ✅

PESTAÑAS QUE SE LLENAN SOLO CUANDO HAY CAMBIO:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Al finalizar el recorrido el bot pregunta:
  ¿Hubo cambio de MANGAS en esta inspección?
    SÍ → solicita datos de cada manga nueva
  ¿Hubo cambio en HILOS del ODF?
    SÍ → solicita posición ODF y estado de hilos
"""


# ══════════════════════════════════════════════════════════════════════════════
#  NOMBRE DEL ARCHIVO FINAL
# ══════════════════════════════════════════════════════════════════════════════

def nombre_archivo(datos: dict) -> str:
    r    = datos["recorrido"]
    ruta = r["nombre_ruta"].split()[0].replace("/", "-")
    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    return f"FOR_FO_02_{ruta}_{fecha}.xlsx"
