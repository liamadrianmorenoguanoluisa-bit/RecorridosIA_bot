import os, io, hmac, struct, time, base64, hashlib, json, logging, threading
import httpx
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, ConversationHandler, ContextTypes, filters, CallbackQueryHandler

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BOT_TOKEN       = os.getenv("BOT_TOKEN")
GEMINI_API_KEY  = os.getenv("GEMINI_API_KEY")
MAPILLARY_TOKEN = os.getenv("MAPILLARY_TOKEN")
TOTP_SECRET     = os.getenv("TOTP_SECRET")
DOMINIO         = os.getenv("DOMINIO_EMAIL", "telconet.ec")
GEMINI_URL      = "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key=" + (GEMINI_API_KEY or "")

USUARIOS_AUTENTICADOS = set()
RUTAS_GUARDADAS = {}

(ESPERANDO_TOTP, MENU_PRINCIPAL, NOMBRE_RUTA, CODIGO_CUADRILLA, NODO_INICIAL, NODO_FINAL,
 LIDER, AYUDANTE, COORDINADOR, PLACA, DISTANCIA,
 CIU_HERRAMIENTAS, CIU_EQUIPOS, CIU_MATERIALES,
 NOVEDADES_AUTO, TAREA_PENDIENTE, FOTO_ANTES, FOTO_DESPUES, OBSERVACIONES,
 MPRIU_CHECK, PREGUNTA_MANGAS, PREGUNTA_HILOS,
 MANGA_NOMBRE, MANGA_COORDS, MANGA_OBS, HILO_ODF, HILO_DATOS,
 NUEVA_RUTA_NOMBRE, NUEVA_RUTA_VIDEO,
 GENERAR_CONFIRMAR, GENERAR_NOMBRE_RUTA, GENERAR_CUADRILLA,
 GENERAR_NODO_INI, GENERAR_NODO_FIN, GENERAR_LIDER,
 GENERAR_AYUDANTE, GENERAR_COORDINADOR, GENERAR_PLACA,
 GENERAR_DISTANCIA, GENERAR_FECHA, GENERAR_HORA_INI,
 GENERAR_HORA_FIN, GENERAR_NOVEDADES,
 TAB_MENU, TAB_CIU_HERR, TAB_CIU_EQUI, TAB_CIU_MATE,
 TAB_MPRIU, TAB_REPORTES, TAB_NOVEDADES_IA,
 VIDEO_BASE_NOMBRE, VIDEO_BASE_UPLOAD) = range(52)

NOVEDADES_MPRIU = [
    "HERRAJES EN MAL ESTADO.", "FALTA DE HERRAJES.", "POSTES EN MAL ESTADO.",
    "POSTE(S) CAMBIADO(S).", "POSTES POR INSTALAR.", "POSTE NUEVO INSTALADO - TN.",
    "POSTE NUEVO INSTALADO - EMPRESAS ELECTRICAS.", "POSTES INCLINADOS.",
    "RETENIDA(S) EN MAL ESTADO.", "RETENIDA(S) CORTADA(S).", "VANOS POR RETEMPLAR.",
    "MANGAS SUELTAS.", "MANGAS ABIERTAS/DANADAS.", "RESERVAS SUELTAS.",
    "CRUCES DE VIAS BAJOS.", "VEGETACION SOBRE FIBRA/MANGA.", "LOCALIZACION DE MANGA.",
    "DOCUMENTACION UNIFILAR DE HILOS.", "LINEA ELECTRICA EN MAL ESTADO.",
    "REGENERACION URBANA.", "AMPLIACION DE VIA.", "CABLE LASTIMADO.",
    "FIBRA INSTALADA INCORRECTAMENTE SOBRE MORDAZA.", "POZO SIN TAPA O EN MAL ESTADO.",
    "REPINTADO DE POZO.", "REPINTADO DE POSTE.", "ELEMENTOS SIN ETIQUETAS ACRILICAS.",
    "RIESGO DE DERRUMBE O DESLAVE.", "RIESGO DE INUNDACIONES.", "RIESGO DE INCENDIO.",
    "NO SE REGISTRAN NOVEDADES DURANTE LA INSPECCION.",
]

SIN_NOV_MOTIVO  = "NO SE REGISTRAN NOVEDADES DURANTE LA INSPECCION."
SIN_NOV_REMEDIO = "NO SE ENCUENTRAN NOVEDADES QUE SIGNIFIQUEN RIESGOS EN EL CABLE DE LA RED INTERURBANO."

REMEDIOS = {
    "VEGETACION SOBRE FIBRA/MANGA.": "REALIZAR LA PODA O RETIRO DE VEGETACION QUE COMPROMETA LA INTEGRIDAD DEL CABLE.",
    "HERRAJES EN MAL ESTADO.": "REALIZAR EL REEMPLAZO INMEDIATO DEL HERRAJE AFECTADO.",
    "POSTES EN MAL ESTADO.": "DOCUMENTAR Y REPORTAR PARA GESTIONAR EL REEMPLAZO DEL POSTE.",
    "POSTES INCLINADOS.": "DOCUMENTAR Y REPORTAR PARA GESTIONAR EL APLOME DEL POSTE.",
    "MANGAS SUELTAS.": "ASEGURAR LA MANGA AL POSTE EN CONFIGURACION TIPO FIGURA 8.",
    "MANGAS ABIERTAS/DANADAS.": "REEMPLAZAR TAPAS Y SELLOS GARANTIZANDO EL CIERRE HERMETICO.",
    "CABLE LASTIMADO.": "DOCUMENTAR E INFORMAR PARA PROGRAMAR EL CAMBIO DEL TRAMO.",
    "DOCUMENTACION UNIFILAR DE HILOS.": "DOCUMENTAR O SOLICITAR PROGRAMACION; UTILIZAR SEGUIDOR DE SENAL.",
    "CRUCES DE VIAS BAJOS.": "AJUSTAR LA ALTURA DEL CABLE A LA DISTANCIA REGLAMENTARIA.",
    "POZO SIN TAPA O EN MAL ESTADO.": "SOLICITAR TRABAJOS DE OBRA CIVIL PARA SU CORRECCION.",
    "FALTA DE HERRAJES.": "INSTALAR HERRAJES CONFORME A LA NORMATIVA TECNICA.",
    "VANOS POR RETEMPLAR.": "REALIZAR EL RETEMPLADO DEL CABLE PARA RESTABLECER LA TENSION.",
    "RESERVAS SUELTAS.": "REORGANIZAR Y ASEGURAR LA RESERVA EN FIGURA 8.",
    "ELEMENTOS SIN ETIQUETAS ACRILICAS.": "VERIFICAR Y COLOCAR ETIQUETA ACRILICA CON EL CODIGO DE RUTA.",
    "RIESGO DE DERRUMBE O DESLAVE.": "DOCUMENTAR Y SOLICITAR REUBICACION DEL RECORRIDO DEL CABLE.",
    "RIESGO DE INUNDACIONES.": "DOCUMENTAR Y SOLICITAR REUBICACION DEL RECORRIDO DEL CABLE.",
    "RIESGO DE INCENDIO.": "DOCUMENTAR Y SOLICITAR REUBICACION DEL RECORRIDO DEL CABLE.",
    "LINEA ELECTRICA EN MAL ESTADO.": "DOCUMENTAR Y SOLICITAR REPORTE AL AREA DE REGULATORIO.",
    "AMPLIACION DE VIA.": "DOCUMENTAR Y COORDINAR MEDIDAS DE MITIGACION CON EL COORDINADOR DE FO.",
}

HERR = ["Cinturon y Linea de Vida","Casco","Escalera de 24 pies","Escalera de 28 pies","Escalera de 32 pies","Conos reflectivos","Caja para herramientas","Juego de destornilladores","Martillo mediano","Estiletes","Cortafrio","Alicate","Llave francesa","Juego de rachet","Pares de guantes aislantes","Tecle","Machete","Cizalla","Pata de cabra","Flejadora (Maquina Eriband)","Extension con foco","Motosierra","Tijeras metalicas","Arco de sierra","Binoculares","Parasol","Remolque / Carrete para F.O."]
EQUI = ["Fusionadora","Cortadora de fibra","Bobina de lanzamiento","OTDR con cargador","Llave Acsys","GPS","Inversor","Etiquetadora"]
MATE = ["Fibra 48h (500mt)","Mangas de 48h y/o 144h (2 minimo)","Rollo de cinta Eriband 3/4","Hebillas para cinta Eriband 3/4","Hojas de sierra","Patchcord de fibra","Adaptadores (Simplex-Duplex)","Paquetes de amarras","Mesas plasticas","Sillas plasticas","Cuchillos","Poleas","Sogas de nylon medianas","Sogas de nylon gruesas","Repelente contra insectos","Repelente contra abejas y avispas"]

AZUL="0000FF"; GRIS="969696"; GRIS2="D9D9D9"; GRIS3="C0C0C0"
AZUL2="0070C0"; VERDE="00B050"; ROJO="FF0000"; BLANCO="FFFFFF"

# ── TOTP ──────────────────────────────────────────────────────────────────────
def verificar_totp(codigo):
    if not TOTP_SECRET:
        return True
    try:
        secreto = base64.b32decode(TOTP_SECRET.upper().replace(" ",""), casefold=True)
        ahora = int(time.time()) // 30
        for delta in [0, -1, 1]:
            contador = struct.pack(">Q", ahora + delta)
            mac = hmac.new(secreto, contador, hashlib.sha1).digest()
            offset = mac[-1] & 0x0F
            c = struct.unpack(">I", mac[offset:offset+4])[0] & 0x7FFFFFFF
            if str(c % 1000000).zfill(6) == str(codigo).strip():
                return True
    except Exception:
        pass
    return False

# ── GEMINI ────────────────────────────────────────────────────────────────────
async def analizar_imagen(img_bytes):
    img_b64 = base64.b64encode(img_bytes).decode()
    prompt = "Analiza esta imagen de inspeccion de fibra optica. Si hay problema responde JSON: {\"tiene_novedad\": true, \"motivo\": \"NOMBRE EN MAYUSCULAS\", \"coordenadas\": \"\"}. Si todo bien: {\"tiene_novedad\": false, \"motivo\": \"\", \"coordenadas\": \"\"}. Solo JSON."
    payload = {"contents": [{"parts": [{"text": prompt}, {"inline_data": {"mime_type": "image/jpeg", "data": img_b64}}]}]}
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(GEMINI_URL, json=payload)
            texto = resp.json()["candidates"][0]["content"]["parts"][0]["text"]
            texto = texto.replace("```json","").replace("```","").strip()
            r = json.loads(texto)
            if not r.get("tiene_novedad"):
                return None
            motivo = r.get("motivo","").upper()
            return {"motivo": motivo, "remedio": REMEDIOS.get(motivo, "DOCUMENTAR Y REPORTAR AL COORDINADOR."), "coordenadas": r.get("coordenadas",""), "tarea_pendiente": "", "foto_antes": img_bytes, "foto_despues": None}
    except Exception as e:
        logger.error("Gemini error: " + str(e))
        return None

# ── EXCEL ─────────────────────────────────────────────────────────────────────
def cel(ws, f, c, v, bold=False, bg=None, halign="left", color="000000", merge_end=None):
    try:
        cell = ws.cell(f, c, v)
        cell.font = Font(bold=bold, name="Calibri", size=11, color=color)
        cell.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=True)
        if bg: cell.fill = PatternFill("solid", fgColor=bg)
        if merge_end: ws.merge_cells(start_row=f, start_column=c, end_row=merge_end[0], end_column=merge_end[1])
        return cell
    except Exception as e:
        logger.warning("cel error: " + str(e))
        return ws.cell(f, c, str(v) if v else "")

def generar_excel(datos):
    r   = datos["recorrido"]
    ciu = datos["ciu"]
    nch = datos["mpriu"].get("novedades_check", {})
    wb  = Workbook()

    # HOJA 1
    ws1 = wb.active; ws1.title = "REPORTES_DE_RECORRIDOS"
    ws1.column_dimensions["A"].width = 41; ws1.column_dimensions["B"].width = 35
    ws1.column_dimensions["C"].width = 32; ws1.column_dimensions["D"].width = 30
    ws1.row_dimensions[2].height = 57
    cel(ws1,2,2,"REPORTE DE RECORRIDOS DE MANTENIMIENTO PREVENTIVO PARA RUTAS INTERURBANAS",bold=True,bg=AZUL,color=BLANCO,halign="center",merge_end=(2,3))
    cel(ws1,2,4,"Codigo: FOR FO 02 Version: 3 (28/05/2021)",bold=True)
    ws1.row_dimensions[4].height = 24
    cel(ws1,4,1,"REPORTE DE RECORRIDO DE RUTAS INTERURBANAS DE F. O.",bold=True,bg=AZUL,color=BLANCO,halign="center",merge_end=(4,4))
    ws1.row_dimensions[5].height = 38; ws1.row_dimensions[6].height = 38
    ws1.merge_cells(start_row=5,start_column=1,end_row=6,end_column=1)
    cel(ws1,5,1,"FECHA Y HORA DEL RECORRIDO",bold=True,bg=GRIS,halign="center")
    cel(ws1,5,2,"FECHA",bold=True); cel(ws1,5,3,"HORA INICIO",bold=True); cel(ws1,5,4,"HORA FIN",bold=True)
    cel(ws1,6,2,r.get("fecha","")); cel(ws1,6,3,r.get("hora_inicio","")); cel(ws1,6,4,r.get("hora_fin",""))
    for i,(label,valor) in enumerate([("NOMBRE DE LA RUTA",r.get("nombre_ruta","")),("CODIGO DE CUADRILLA",r.get("codigo_cuadrilla","")),("NODO INICIAL",r.get("nodo_inicial",""))]):
        f=7+i; ws1.row_dimensions[f].height=38
        cel(ws1,f,1,label,bold=True,bg=GRIS2); cel(ws1,f,2,valor,bold=True,merge_end=(f,4))
    fila=10
    for nov in r.get("novedades",[]):
        num=str(nov.get("numero",""))
        ws1.row_dimensions[fila].height=38
        ws1.merge_cells(start_row=fila,start_column=1,end_row=fila+1,end_column=1)
        cel(ws1,fila,1,"FECHA Y HORA NOVEDAD # "+num,bold=True,bg=GRIS,halign="center")
        cel(ws1,fila,2,"FECHA",bold=True); cel(ws1,fila,3,"HORA INICIO",bold=True); cel(ws1,fila,4,"HORA FIN",bold=True)
        fila+=1; ws1.row_dimensions[fila].height=38
        cel(ws1,fila,2,nov.get("fecha","")); cel(ws1,fila,3,nov.get("hora_inicio","")); cel(ws1,fila,4,nov.get("hora_fin",""))
        fila+=1
        for label,key in [("MOTIVO APARENTE DE LA NOVEDAD","motivo"),("REMEDIO DEFINITIVO A LA NOVEDAD","remedio"),("TAREA PENDIENTE","tarea_pendiente"),("COORDENADAS","coordenadas")]:
            ws1.row_dimensions[fila].height=42
            cel(ws1,fila,1,label,bold=True,bg=GRIS3); cel(ws1,fila,2,nov.get(key,""),merge_end=(fila,4)); fila+=1
    for label,valor in [("NODO FINAL",r.get("nodo_final","")),("LIDER DE CUADRILLA QUE ELABORA INFORME",r.get("lider","")),("AYUDANTE TECNICO",r.get("ayudante","")),("COORDINADOR FIBRA OPTICA",r.get("coordinador","")),("FOTOS ANEXAS AL REPORTE",str(r.get("fotos_total",0))),("OBSERVACIONES GENERALES",r.get("observaciones",""))]:
        ws1.row_dimensions[fila].height=38
        cel(ws1,fila,1,label,bold=True,bg=GRIS2); cel(ws1,fila,2,valor,merge_end=(fila,4)); fila+=1

    # HOJA 2
    ws2=wb.create_sheet("FOTOS_ANEXAS_AL_REPORTE")
    ws2.column_dimensions["B"].width=20; ws2.column_dimensions["C"].width=59; ws2.column_dimensions["D"].width=24
    cel(ws2,2,3,"REPORTE DE RECORRIDOS DE MANTENIMIENTO PREVENTIVO PARA RUTAS INTERURBANAS",bold=True,bg=AZUL,color=BLANCO,halign="center",merge_end=(2,4))
    cel(ws2,4,2,"FOTOS DE LAS ACCIONES CORRECTIVAS",bold=True,bg=AZUL,color=BLANCO,halign="center",merge_end=(4,5))
    cel(ws2,7,2,"NODO INICIO RECORRIDO",bold=True); cel(ws2,7,3,"FOTO",bold=True,bg=AZUL,color=BLANCO,halign="center"); cel(ws2,7,4,"NOMBRE DEL NODO",bold=True,bg=AZUL,color=BLANCO,halign="center",merge_end=(7,5))
    ws2.row_dimensions[8].height=315; cel(ws2,8,3,"[FOTO NODO INICIO]"); cel(ws2,8,4,r.get("nodo_inicial",""),bold=True,merge_end=(8,5))
    f2=10
    for nov in r.get("novedades",[]):
        cel(ws2,f2,2,"NOVEDAD # "+str(nov.get("numero","")),bold=True)
        cel(ws2,f2,3,"ANTES DEL MANTENIMIENTO",bold=True,bg=AZUL,color=BLANCO,halign="center")
        cel(ws2,f2,4,"DESPUES DEL MANTENIMIENTO",bold=True,bg=AZUL,color=BLANCO,halign="center",merge_end=(f2,5))
        f2+=1; ws2.row_dimensions[f2].height=315
        cel(ws2,f2,3,"[FOTO ANTES]"); cel(ws2,f2,4,"[FOTO DESPUES]",merge_end=(f2,5)); f2+=2
    cel(ws2,f2,2,"NODO FINAL DEL RECORRIDO",bold=True); cel(ws2,f2,3,"FOTO",bold=True,bg=AZUL,color=BLANCO,halign="center"); cel(ws2,f2,4,r.get("nodo_final",""),bold=True,merge_end=(f2,5))

    # HOJA 3
    ws3=wb.create_sheet("MANGAS")
    ws3.column_dimensions["B"].width=25; ws3.column_dimensions["C"].width=35; ws3.column_dimensions["D"].width=25; ws3.column_dimensions["E"].width=35
    cel(ws3,2,3,"REPORTE DE RECORRIDOS",bold=True,bg=AZUL,color=BLANCO,halign="center",merge_end=(2,4))
    cel(ws3,4,2,"FOTOS DE LAS MANGAS DESDE EL NODO A AL B",bold=True,bg="00133A",color=BLANCO,halign="center",merge_end=(4,5))
    mangas=datos.get("mangas",[]); f3=6
    if not mangas: cel(ws3,f3,2,"SIN CAMBIO DE MANGAS EN ESTE RECORRIDO")
    else:
        for i in range(0,len(mangas),2):
            m1=mangas[i]; m2=mangas[i+1] if i+1<len(mangas) else {}
            ws3.row_dimensions[f3].height=315; cel(ws3,f3,3,"[FOTO MANGA]"); cel(ws3,f3,5,"[FOTO MANGA]"); f3+=1
            for label,k in [("NOMBRE:","nombre"),("DERIVACION:","derivacion"),("COORDENADAS:","coordenadas"),("OBSERVACION:","observacion")]:
                cel(ws3,f3,2,label,bold=True,bg="1F4E79",color=BLANCO); cel(ws3,f3,3,m1.get(k,""))
                cel(ws3,f3,4,label,bold=True,bg="1F4E79",color=BLANCO); cel(ws3,f3,5,m2.get(k,"")); f3+=1

    # HOJA 4
    ws4=wb.create_sheet("INVENTARIO DE HILOS EN NODO")
    ws4.column_dimensions["A"].width=10; ws4.column_dimensions["B"].width=14; ws4.column_dimensions["C"].width=45
    cel(ws4,2,3,"REPORTE DE RECORRIDOS",bold=True,bg=AZUL,color=BLANCO,halign="center",merge_end=(2,6))
    cel(ws4,3,1,"NODO: ",bold=True); cel(ws4,3,3,r.get("nodo_final",""))
    cel(ws4,4,1,"NOMBRE ODF DE RUTA:",bold=True); cel(ws4,4,3,datos["hilos"].get("posicion_odf",""))
    for col,txt in [(2,"PAR"),(3,"HILO"),(4,"NOMENCLATURA"),(5,"RACK #")]:
        cel(ws4,6,col,txt,bold=True,bg=AZUL2,color=BLANCO,halign="center")
    f4=7; hilos=datos["hilos"].get("filas",[])
    if not hilos: cel(ws4,f4,2,"SIN CAMBIOS EN ODF EN ESTE RECORRIDO")
    else:
        for h in hilos:
            cel(ws4,f4,2,h.get("hilo_par","")); cel(ws4,f4,4,h.get("descripcion","")); cel(ws4,f4,5,h.get("estado","")); f4+=1

    # HOJA 5
    ws5=wb.create_sheet("Checklist CIU")
    for col,w in [("A",9),("B",26),("C",11),("D",21),("E",11),("F",14),("G",11),("H",14)]: ws5.column_dimensions[col].width=w
    ws5.row_dimensions[2].height=46
    cel(ws5,2,2,"CHECKLIST CUADRILLA INTERURBANA",bold=True,bg=AZUL,color=BLANCO,halign="center",merge_end=(2,7))
    cel(ws5,2,8,"Codigo: FOR FO 05",bold=True)
    cel(ws5,3,2,"Fecha del Recorrido",bold=True); cel(ws5,3,3,r.get("fecha",""))
    cel(ws5,3,5,"Hora Inicio",bold=True); cel(ws5,3,6,r.get("hora_inicio",""))
    cel(ws5,3,7,"Hora Fin",bold=True); cel(ws5,3,8,r.get("hora_fin",""))
    f5=4
    for label,valor in [("Nombre de Ruta",r.get("nombre_ruta","")),("Nodo Inicio",r.get("nodo_inicial","")),("Nodo Final",r.get("nodo_final","")),("Distancia",ciu.get("distancia_ruta","")),("Lider de Cuadrilla",r.get("lider","")),("Vehiculo Placa",ciu.get("vehiculo_placa","")),("Coordinador",r.get("coordinador",""))]:
        cel(ws5,f5,2,label,bold=True); cel(ws5,f5,3,valor,merge_end=(f5,8)); f5+=1
    for sec,items,data_s in [("HERRAMIENTAS Y EPP",HERR,ciu.get("herramientas",{})),("EQUIPOS ELECTRONICOS",EQUI,ciu.get("equipos",{})),("MATERIALES E INSUMOS",MATE,ciu.get("materiales",{}))]:
        cel(ws5,f5,2,sec,bold=True,bg=AZUL2,color=BLANCO,halign="center",merge_end=(f5,3))
        cel(ws5,f5,4,"CANTIDAD",bold=True,bg=AZUL2,color=BLANCO,halign="center")
        cel(ws5,f5,5,"OBSERVACIONES",bold=True,bg=AZUL2,color=BLANCO,halign="center",merge_end=(f5,8)); f5+=1
        for nombre in items:
            info=data_s.get(nombre,{})
            cant=info.get("cantidad",0) if isinstance(info,dict) else int(info or 0)
            obs=info.get("obs","NINGUNA") if isinstance(info,dict) else ("BUEN ESTADO" if cant>0 else "NINGUNA")
            cel(ws5,f5,2,nombre); cel(ws5,f5,4,cant,halign="center")
            bg_o=VERDE if obs=="BUEN ESTADO" else (ROJO if obs=="MAL ESTADO" else "808080")
            cel(ws5,f5,5,obs,bg=bg_o,color=BLANCO,halign="center",merge_end=(f5,8)); f5+=1

    # HOJA 6
    ws6=wb.create_sheet("Checklists MPRIU")
    for col,w in [("A",9),("B",26),("C",11),("D",32),("E",11),("F",14),("G",11),("H",14)]: ws6.column_dimensions[col].width=w
    ws6.row_dimensions[2].height=46
    cel(ws6,2,2,"CHECKLIST DE RECORRIDO DE MANTENIMIENTO PREVENTIVO DE RUTAS INTERURBANAS",bold=True,bg=AZUL,color=BLANCO,halign="center",merge_end=(2,7))
    cel(ws6,2,8,"Codigo: FOR FO 08",bold=True)
    cel(ws6,3,2,"Fecha del Recorrido",bold=True); cel(ws6,3,3,r.get("fecha",""))
    cel(ws6,3,5,"Hora Inicio",bold=True); cel(ws6,3,6,r.get("hora_inicio",""))
    cel(ws6,3,7,"Hora Fin",bold=True); cel(ws6,3,8,r.get("hora_fin",""))
    f6=4
    for label,valor in [("Nombre de Ruta",r.get("nombre_ruta","")),("Nodo Inicio",r.get("nodo_inicial","")),("Nodo Final",r.get("nodo_final","")),("Distancia",ciu.get("distancia_ruta","")),("Lider de Cuadrilla",r.get("lider","")),("Vehiculo Placa",ciu.get("vehiculo_placa","")),("Coordinador",r.get("coordinador",""))]:
        cel(ws6,f6,2,label,bold=True); cel(ws6,f6,3,valor,merge_end=(f6,8)); f6+=1
    f6+=1
    cel(ws6,f6,2,"NOVEDAD",bold=True,bg=AZUL2,color=BLANCO,halign="center")
    cel(ws6,f6,3,"CHECK",bold=True,bg=AZUL2,color=BLANCO,halign="center")
    cel(ws6,f6,4,"SOLUCION",bold=True,bg=AZUL2,color=BLANCO,halign="center",merge_end=(f6,7))
    cel(ws6,f6,8,"CANTIDAD",bold=True,bg=AZUL2,color=BLANCO,halign="center"); f6+=1
    SOLUCIONES={"HERRAJES EN MAL ESTADO.":"REALIZAR EL REEMPLAZO INMEDIATO DEL HERRAJE AFECTADO.","FALTA DE HERRAJES.":"INSTALAR LOS HERRAJES CONFORME A LA NORMATIVA TECNICA.","POSTES EN MAL ESTADO.":"DOCUMENTAR Y REPORTAR PARA GESTIONAR EL REEMPLAZO DEL POSTE.","POSTES INCLINADOS.":"DOCUMENTAR Y REPORTAR PARA GESTIONAR EL APLOME DEL POSTE.","VANOS POR RETEMPLAR.":"REALIZAR EL RETEMPLADO DEL CABLE PARA RESTABLECER LA TENSION.","MANGAS SUELTAS.":"ASEGURAR LA MANGA AL POSTE EN CONFIGURACION TIPO FIGURA 8.","MANGAS ABIERTAS/DANADAS.":"REEMPLAZAR TAPAS Y SELLOS GARANTIZANDO EL CIERRE HERMETICO.","RESERVAS SUELTAS.":"REORGANIZAR Y ASEGURAR LA RESERVA EN FIGURA 8.","CRUCES DE VIAS BAJOS.":"AJUSTAR LA ALTURA DEL CABLE A LA DISTANCIA REGLAMENTARIA.","VEGETACION SOBRE FIBRA/MANGA.":"REALIZAR LA PODA O RETIRO DE VEGETACION QUE COMPROMETA LA INTEGRIDAD DEL CABLE.","DOCUMENTACION UNIFILAR DE HILOS.":"DOCUMENTAR O SOLICITAR PROGRAMACION; UTILIZAR SEGUIDOR DE SENAL.","LINEA ELECTRICA EN MAL ESTADO.":"DOCUMENTAR Y SOLICITAR REPORTE AL AREA DE REGULATORIO.","CABLE LASTIMADO.":"DOCUMENTAR E INFORMAR PARA PROGRAMAR EL CAMBIO DEL TRAMO.","POZO SIN TAPA O EN MAL ESTADO.":"SOLICITAR TRABAJOS DE OBRA CIVIL PARA SU CORRECCION.","ELEMENTOS SIN ETIQUETAS ACRILICAS.":"VERIFICAR Y COLOCAR ETIQUETA ACRILICA CON EL CODIGO DE RUTA.","RIESGO DE DERRUMBE O DESLAVE.":"DOCUMENTAR Y SOLICITAR REUBICACION DEL RECORRIDO DEL CABLE.","RIESGO DE INUNDACIONES.":"DOCUMENTAR Y SOLICITAR REUBICACION DEL RECORRIDO DEL CABLE.","RIESGO DE INCENDIO.":"DOCUMENTAR Y SOLICITAR REUBICACION DEL RECORRIDO DEL CABLE.","NO SE REGISTRAN NOVEDADES DURANTE LA INSPECCION.":"NO SE ENCUENTRAN NOVEDADES QUE SIGNIFIQUEN RIESGOS EN EL CABLE DE LA RED INTERURBANO."}
    for novedad in NOVEDADES_MPRIU:
        ws6.row_dimensions[f6].height=43
        info=nch.get(novedad,{}); tiene=info.get("check",False); cant=info.get("cantidad",0); chk="SI" if tiene else "NO"
        sol=SOLUCIONES.get(novedad,"DOCUMENTAR Y REPORTAR AL COORDINADOR.")
        cel(ws6,f6,2,novedad); cel(ws6,f6,3,chk,bold=True,bg=(VERDE if tiene else ROJO),color=BLANCO,halign="center")
        cel(ws6,f6,4,sol,merge_end=(f6,7)); cel(ws6,f6,8,cant if tiene else 0,halign="center"); f6+=1
    ws6.row_dimensions[f6].height=60
    cel(ws6,f6,2,"Observaciones:",bold=True); cel(ws6,f6,3,r.get("observaciones",""),merge_end=(f6,8))

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

def datos_vacios():
    return {
        "recorrido":{"fecha":datetime.now().strftime("%d/%m/%Y"),"hora_inicio":"","hora_fin":"","nombre_ruta":"","codigo_cuadrilla":"","nodo_inicial":"","nodo_final":"","lider":"","ayudante":"","coordinador":"","fotos_total":0,"observaciones":"","novedades":[]},
        "ciu":{"vehiculo_placa":"","distancia_ruta":"","herramientas":{},"equipos":{},"materiales":{}},
        "mpriu":{"novedades_check":{},"observaciones":""},
        "mangas":[],"hilos":{"posicion_odf":"","filas":[]},
    }

def novedad_vacia(numero):
    ahora=datetime.now()
    return {"numero":numero,"fecha":ahora.strftime("%d/%m/%Y"),"hora_inicio":ahora.strftime("%H:%M:%S"),"hora_fin":ahora.strftime("%H:%M:%S"),"motivo":"","remedio":"","tarea_pendiente":"","coordenadas":"","foto_antes":None,"foto_despues":None}

def nombre_archivo(datos):
    ruta=datos["recorrido"]["nombre_ruta"].split()[0].replace("/","-") if datos["recorrido"]["nombre_ruta"] else "RUTA"
    return "FOR_FO_02_"+ruta+"_"+datetime.now().strftime("%Y%m%d_%H%M")+".xlsx"

# ── AUTENTICACION ─────────────────────────────────────────────────────────────
async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id in USUARIOS_AUTENTICADOS:
        return await menu_principal(update, ctx)
    await update.message.reply_text(
        "RecorridosIA - Acceso restringido" + chr(10) + chr(10) +
        "Ingresa tu correo y codigo de 6 digitos:" + chr(10) + chr(10) +
        "email: tucorreo@telconet.ec" + chr(10) + "totp: 123456",
        reply_markup=ReplyKeyboardRemove()
    )
    return ESPERANDO_TOTP

async def handler_totp(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    texto=update.message.text.strip().lower(); email=""; codigo=""
    for linea in texto.splitlines():
        if linea.startswith("email:"): email=linea.replace("email:","").strip()
        elif linea.startswith("totp:"): codigo=linea.replace("totp:","").strip()
    if not email or not codigo:
        await update.message.reply_text("Formato incorrecto. Usa:" + chr(10) + "email: tucorreo@telconet.ec" + chr(10) + "totp: 123456")
        return ESPERANDO_TOTP
    if not email.endswith(DOMINIO):
        await update.message.reply_text("Solo correos @"+DOMINIO)
        return ESPERANDO_TOTP
    if verificar_totp(codigo):
        USUARIOS_AUTENTICADOS.add(update.effective_user.id)
        nombre=email.split("@")[0].upper()
        await update.message.reply_text("Acceso autorizado. Bienvenido "+nombre)
        return await menu_principal(update, ctx)
    await update.message.reply_text("Codigo incorrecto." + chr(10) + "email: tucorreo@telconet.ec" + chr(10) + "totp: 123456")
    return ESPERANDO_TOTP

async def menu_principal(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    teclado=[["Generar Informe","Nueva Ruta Base"],["Mis Rutas","Ayuda"]]
    await update.message.reply_text("RecorridosIA - Menu principal",reply_markup=ReplyKeyboardMarkup(teclado,resize_keyboard=True))
    return MENU_PRINCIPAL

# ── MENU DE PESTANAS ──────────────────────────────────────────────────────────
async def generar_informe(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in USUARIOS_AUTENTICADOS:
        return await start(update, ctx)
    if "datos" not in ctx.user_data:
        ctx.user_data["datos"] = datos_vacios()
    return await tab_menu(update, ctx)

async def tab_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in USUARIOS_AUTENTICADOS:
        return await start(update, ctx)
    if "datos" not in ctx.user_data:
        ctx.user_data["datos"] = datos_vacios()
    datos=ctx.user_data["datos"]; r=datos["recorrido"]
    ciu_ok=bool(datos["ciu"].get("vehiculo_placa"))
    mpriu_ok=bool(datos["mpriu"].get("novedades_check"))
    rep_ok=bool(r.get("nombre_ruta"))
    novedades=len(r.get("novedades",[]))
    man_ok=bool(datos.get("mangas"))
    hil_ok=bool(datos["hilos"].get("filas"))
    def v(ok): return "✅" if ok else "⬜"
    teclado=InlineKeyboardMarkup([
        [InlineKeyboardButton(v(ciu_ok)+" Checklist CIU",callback_data="tab_1")],
        [InlineKeyboardButton(v(mpriu_ok)+" Checklists MPRIU",callback_data="tab_2")],
        [InlineKeyboardButton(v(rep_ok)+" REPORTES_DE_RECORRIDOS",callback_data="tab_reportes")],
        [InlineKeyboardButton(("✅" if novedades>0 else "⬜")+" FOTOS_ANEXAS_AL_REPORTE ["+str(novedades)+" nov]",callback_data="tab_fotos")],
        [InlineKeyboardButton(v(man_ok)+" Mangas",callback_data="tab_5"),InlineKeyboardButton(v(hil_ok)+" Hilos ODF",callback_data="tab_6")],
        [InlineKeyboardButton("GENERAR EXCEL",callback_data="tab_generar")],
    ])
    completadas=sum([ciu_ok,mpriu_ok,rep_ok,novedades>0])
    msg="INFORME FOR FO 02 — Completado: "+str(completadas)+"/4"+chr(10)+"Selecciona la pestana:"
    if update.callback_query:
        await update.callback_query.answer()
        try: await update.callback_query.edit_message_text(msg,reply_markup=teclado)
        except: await update.callback_query.message.reply_text(msg,reply_markup=teclado)
    else:
        await update.message.reply_text(msg,reply_markup=teclado)
    return TAB_MENU

async def tab_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query=update.callback_query; await query.answer(); data=query.data
    volver=InlineKeyboardMarkup([[InlineKeyboardButton("Volver al menu",callback_data="tab_menu")]])

    if data=="tab_generar":
        return await enviar_excel(update, ctx)

    elif data=="tab_menu":
        return await tab_menu(update, ctx)

    elif data=="tab_1":
        msg=("CHECKLIST CIU - HERRAMIENTAS Y EPP"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+
             "Escribe cantidades separadas por coma:"+chr(10)+chr(10)+
             " 1. Cinturon y Linea de Vida"+chr(10)+" 2. Casco"+chr(10)+" 3. Escalera 24 pies"+chr(10)+
             " 4. Escalera 28 pies"+chr(10)+" 5. Escalera 32 pies"+chr(10)+" 6. Conos reflectivos"+chr(10)+
             " 7. Caja para herramientas"+chr(10)+" 8. Juego destornilladores"+chr(10)+" 9. Martillo mediano"+chr(10)+
             "10. Estiletes"+chr(10)+"11. Cortafrio"+chr(10)+"12. Alicate"+chr(10)+"13. Llave francesa"+chr(10)+
             "14. Juego de rachet"+chr(10)+"15. Guantes aislantes (pares)"+chr(10)+"16. Tecle"+chr(10)+
             "17. Machete"+chr(10)+"18. Cizalla"+chr(10)+"19. Pata de cabra"+chr(10)+"20. Flejadora Eriband"+chr(10)+
             "21. Extension con foco"+chr(10)+"22. Motosierra"+chr(10)+"23. Tijeras metalicas"+chr(10)+
             "24. Arco de sierra"+chr(10)+"25. Binoculares"+chr(10)+"26. Parasol"+chr(10)+"27. Remolque/Carrete FO"+chr(10)+chr(10)+
             "Ejemplo: 2,2,0,2,0,6,0,1,1,2,2,0,0,1,2,2,2,1,0,0,0,0,0,0,0,0,0")
        await query.edit_message_text(msg,reply_markup=volver)
        ctx.user_data["tab_actual"]="1"; return TAB_CIU_HERR

    elif data=="tab_2":
        msg="CHECKLISTS MPRIU"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+"Escribe los NUMEROS de las novedades (separados por coma):"+chr(10)+chr(10)
        for i,nov in enumerate(NOVEDADES_MPRIU,1): msg+=str(i).rjust(2)+". "+nov+chr(10)
        msg+=chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+"Ejemplo: 16,18"+chr(10)+"Sin novedades: 31"
        await query.edit_message_text(msg,reply_markup=volver)
        ctx.user_data["tab_actual"]="2"; return TAB_MPRIU

    elif data in ("tab_reportes","tab_fotos"):
        nombre_tab="REPORTES_DE_RECORRIDOS" if data=="tab_reportes" else "FOTOS_ANEXAS_AL_REPORTE"
        teclado2=InlineKeyboardMarkup([
            [InlineKeyboardButton("Manual (sin senial)",callback_data="rep_manual"),InlineKeyboardButton("Con IA (Gemini)",callback_data="rep_ia")],
            [InlineKeyboardButton("Volver al menu",callback_data="tab_menu")],
        ])
        await query.edit_message_text(nombre_tab+chr(10)+chr(10)+"Como quieres llenar esta pestana?",reply_markup=teclado2)
        return TAB_MENU

    elif data=="rep_manual":
        msg=("REPORTES - Modo Manual"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+
             "Ingresa los datos (uno por linea):"+chr(10)+chr(10)+
             "RUTA: GOSSEAL-MACHACHI   TAREA: 157415066"+chr(10)+"CUADRILLA: FO UIO INT 04"+chr(10)+
             "NODO_INI: GOSSEAL"+chr(10)+"NODO_FIN: MACHACHI"+chr(10)+"LIDER: RICHARD DAVID TAIPE COYAGO"+chr(10)+
             "AYUDANTE: JOSE LUIS ALLAICA CONDO"+chr(10)+"COORDINADOR: JUAN CARLOS YEPEZ ACAN"+chr(10)+
             "PLACA: PCO3940"+chr(10)+"DISTANCIA: 59KM"+chr(10)+"FECHA: HOY"+chr(10)+
             "HORA_INI: AHORA"+chr(10)+"HORA_FIN: AHORA"+chr(10)+"FOTOS: 6"+chr(10)+"OBS: texto o NINGUNA"+chr(10)+chr(10)+
             "Para novedades agrega:"+chr(10)+"NOV: VEGETACION SOBRE FIBRA/MANGA. | -0.477,-78.579"+chr(10)+chr(10)+
             "Cuando termines escribe: FIN")
        await query.edit_message_text(msg,reply_markup=volver)
        ctx.user_data["tab_actual"]="3"; ctx.user_data["novedades_manuales"]=[]
        return TAB_REPORTES

    elif data=="rep_ia":
        await query.edit_message_text(
            "REPORTES + FOTOS con Gemini IA"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+
            "Envia las fotos de la inspeccion."+chr(10)+"Gemini detectara automaticamente las novedades."+chr(10)+chr(10)+
            "Cuando termines escribe: LISTO",reply_markup=volver)
        ctx.user_data["tab_actual"]="4"; ctx.user_data["media_inspeccion"]=[]
        return TAB_NOVEDADES_IA

    elif data=="tab_5":
        ctx.user_data["datos"]["mangas"]=[]; ctx.user_data["manga_temp"]={}
        await query.edit_message_text(
            "MANGAS"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+
            "Voy a pedirte los datos de cada manga uno por uno."+chr(10)+chr(10)+
            "Nombre de la manga #1:"+chr(10)+"Ejemplo: UIO-B-MAC/GOS-F1-DER-01",reply_markup=volver)
        return MANGA_NOMBRE

    elif data=="tab_6":
        await query.edit_message_text(
            "INVENTARIO DE HILOS EN NODO"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+
            "Posicion del ODF:"+chr(10)+"Ejemplo: MAC-GOS-F01-R04-ODF02-48",reply_markup=volver)
        return HILO_ODF

    elif data in ("manga_der_si","manga_der_no"):
        derivacion="SI" if data=="manga_der_si" else "NO"
        ctx.user_data["manga_temp"]["derivacion"]=derivacion
        await query.edit_message_text(
            "Nombre: "+ctx.user_data["manga_temp"].get("nombre","")+chr(10)+
            "Derivacion: "+derivacion+chr(10)+chr(10)+
            "Coordenadas GPS de la manga:"+chr(10)+"Ejemplo: -0.477057,-78.579350")
        return MANGA_COORDS

    return TAB_MENU

# ── CIU HANDLERS ──────────────────────────────────────────────────────────────
async def tab_ciu_herr(update, ctx):
    valores=[v.strip() for v in update.message.text.replace(","," ").split()]
    herr={}; resumen=""
    for i,nombre in enumerate(HERR):
        cant=int(valores[i]) if i<len(valores) and valores[i].isdigit() else 0
        herr[nombre]={"cantidad":cant,"obs":"BUEN ESTADO" if cant>0 else "NINGUNA"}
        if cant>0: resumen+="  ✅ "+nombre+": "+str(cant)+chr(10)
    ctx.user_data["datos"]["ciu"]["herramientas"]=herr
    await update.message.reply_text(
        "✅ HERRAMIENTAS GUARDADAS"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+
        (resumen if resumen else "  Ninguna")+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+
        "EQUIPOS ELECTRONICOS — escribe cantidades:"+chr(10)+chr(10)+
        " 1. Fusionadora"+chr(10)+" 2. Cortadora de fibra"+chr(10)+" 3. Bobina de lanzamiento"+chr(10)+
        " 4. OTDR con cargador"+chr(10)+" 5. Llave Acsys"+chr(10)+" 6. GPS"+chr(10)+
        " 7. Inversor"+chr(10)+" 8. Etiquetadora"+chr(10)+chr(10)+"Ejemplo: 1,2,0,1,1,0,1,1")
    return TAB_CIU_EQUI

async def tab_ciu_equi(update, ctx):
    valores=[v.strip() for v in update.message.text.replace(","," ").split()]
    equi={}; resumen=""
    for i,nombre in enumerate(EQUI):
        cant=int(valores[i]) if i<len(valores) and valores[i].isdigit() else 0
        equi[nombre]={"cantidad":cant,"obs":"BUEN ESTADO" if cant>0 else "NINGUNA"}
        if cant>0: resumen+="  ✅ "+nombre+": "+str(cant)+chr(10)
    ctx.user_data["datos"]["ciu"]["equipos"]=equi
    await update.message.reply_text(
        "✅ EQUIPOS GUARDADOS"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+
        (resumen if resumen else "  Ninguno")+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+
        "MATERIALES E INSUMOS — escribe cantidades:"+chr(10)+chr(10)+
        " 1. Fibra 48h (500mt)"+chr(10)+" 2. Mangas 48h/144h"+chr(10)+" 3. Rollo cinta Eriband 3/4"+chr(10)+
        " 4. Hebillas Eriband 3/4"+chr(10)+" 5. Hojas de sierra"+chr(10)+" 6. Patchcord de fibra"+chr(10)+
        " 7. Adaptadores (Simplex-Duplex)"+chr(10)+" 8. Paquetes de amarras"+chr(10)+" 9. Mesas plasticas"+chr(10)+
        "10. Sillas plasticas"+chr(10)+"11. Cuchillos"+chr(10)+"12. Poleas"+chr(10)+
        "13. Sogas nylon medianas"+chr(10)+"14. Sogas nylon gruesas"+chr(10)+
        "15. Repelente insectos"+chr(10)+"16. Repelente abejas/avispas"+chr(10)+chr(10)+
        "Ejemplo: 335,2,1,6,0,2,10,2,0,0,0,1,1,0,0,0")
    return TAB_CIU_MATE

async def tab_ciu_mate(update, ctx):
    valores=[v.strip() for v in update.message.text.replace(","," ").split()]
    mate={}; resumen=""
    for i,nombre in enumerate(MATE):
        cant=int(valores[i]) if i<len(valores) and valores[i].isdigit() else 0
        mate[nombre]={"cantidad":cant,"obs":"BUEN ESTADO" if cant>0 else "NINGUNA"}
        if cant>0: resumen+="  ✅ "+nombre+": "+str(cant)+chr(10)
    ctx.user_data["datos"]["ciu"]["materiales"]=mate
    await update.message.reply_text(
        "✅ CHECKLIST CIU COMPLETO"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+
        (resumen if resumen else "  Ninguno")+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    return await tab_menu(update, ctx)

# ── MPRIU HANDLERS ────────────────────────────────────────────────────────────
async def tab_mpriu(update, ctx):
    txt=update.message.text.strip().upper()
    if txt=="OK":
        await update.message.reply_text("✅ CHECKLIST MPRIU GUARDADO!")
        return await tab_menu(update, ctx)
    numeros=[n.strip() for n in txt.replace(","," ").split() if n.strip().isdigit()]
    nch={}
    for num_str in numeros:
        idx=int(num_str)-1
        if 0<=idx<len(NOVEDADES_MPRIU):
            nov=NOVEDADES_MPRIU[idx]
            nch[nov]={"check":True,"cantidad":1}
    ctx.user_data["datos"]["mpriu"]["novedades_check"]=nch
    cant=len([v for v in nch.values() if v.get("check")])
    resumen="".join(["  ✅ "+nov+chr(10) for nov in nch])
    await update.message.reply_text(
        "✅ "+str(cant)+" novedad(es) marcada(s):"+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+
        (resumen if resumen else "  Sin novedades")+chr(10)+"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"+chr(10)+chr(10)+
        "Ajusta cantidades en formato NUMERO:CANTIDAD"+chr(10)+"Ejemplo: 16:5,18:1"+chr(10)+"Si estan correctas escribe: OK")
    return TAB_MPRIU

# ── REPORTES HANDLERS ─────────────────────────────────────────────────────────
async def tab_reportes(update, ctx):
    lineas=update.message.text.strip().split(chr(10))
    datos=ctx.user_data["datos"]; r=datos["recorrido"]
    for linea in lineas:
        if ":" not in linea: continue
        clave,valor=linea.split(":",1); clave=clave.strip().upper(); valor=valor.strip()
        if valor.upper()=="HOY": valor=datetime.now().strftime("%d/%m/%Y")
        elif valor.upper()=="AHORA": valor=datetime.now().strftime("%H:%M:%S")
        if clave=="RUTA": r["nombre_ruta"]=valor.upper()
        elif clave=="CUADRILLA": r["codigo_cuadrilla"]=valor.upper()
        elif clave=="NODO_INI": r["nodo_inicial"]=valor.upper()
        elif clave=="NODO_FIN": r["nodo_final"]=valor.upper()
        elif clave=="LIDER": r["lider"]=valor.upper()
        elif clave=="AYUDANTE": r["ayudante"]=valor.upper()
        elif clave=="COORDINADOR": r["coordinador"]=valor.upper()
        elif clave=="PLACA": datos["ciu"]["vehiculo_placa"]=valor.upper()
        elif clave=="DISTANCIA": datos["ciu"]["distancia_ruta"]=valor.upper()
        elif clave=="FECHA": r["fecha"]=valor
        elif clave=="HORA_INI": r["hora_inicio"]=valor
        elif clave=="HORA_FIN": r["hora_fin"]=valor
        elif clave=="FOTOS": r["fotos_total"]=int(valor) if valor.isdigit() else 0
        elif clave=="OBS":
            if valor.upper()!="NINGUNA": r["observaciones"]=valor.upper()
    novedades_nuevas=[]
    for linea in lineas:
        if linea.upper().startswith("NOV:"):
            partes=linea[4:].strip().split("|")
            motivo=partes[0].strip().upper(); coords=partes[1].strip() if len(partes)>1 else ""
            remedio=REMEDIOS.get(motivo,"DOCUMENTAR Y REPORTAR AL COORDINADOR.")
            nov=novedad_vacia(len(novedades_nuevas)+1)
            nov["motivo"]=motivo; nov["remedio"]=remedio; nov["coordenadas"]=coords
            novedades_nuevas.append(nov)
            datos["mpriu"]["novedades_check"][motivo]={"check":True,"cantidad":1}
        elif linea.upper()=="FIN": break
    if novedades_nuevas: r["novedades"]=novedades_nuevas
    elif not r.get("novedades"):
        nov=novedad_vacia(1); nov["motivo"]=SIN_NOV_MOTIVO; nov["remedio"]=SIN_NOV_REMEDIO
        r["novedades"]=[nov]
    await update.message.reply_text(
        "✅ Datos guardados!"+chr(10)+"Ruta: "+r.get("nombre_ruta","")+chr(10)+
        "Novedades: "+str(len(r.get("novedades",[]))))
    return await tab_menu(update, ctx)

async def tab_novedades_ia(update, ctx):
    if "media_inspeccion" not in ctx.user_data: ctx.user_data["media_inspeccion"]=[]
    if update.message.photo:
        foto=await update.message.photo[-1].get_file()
        ctx.user_data["media_inspeccion"].append(bytes(await foto.download_as_bytearray()))
        await update.message.reply_text("Foto "+str(len(ctx.user_data["media_inspeccion"]))+" recibida. Envia mas o escribe LISTO")
        return TAB_NOVEDADES_IA
    if update.message.text and update.message.text.upper()=="LISTO":
        await update.message.reply_text("Analizando con Gemini IA...")
        media=ctx.user_data.get("media_inspeccion",[]); novedades=[]
        for img in media:
            r_ia=await analizar_imagen(img)
            if r_ia:
                n=novedad_vacia(len(novedades)+1); n.update(r_ia); novedades.append(n)
        if not novedades:
            n=novedad_vacia(1); n["motivo"]=SIN_NOV_MOTIVO; n["remedio"]=SIN_NOV_REMEDIO; novedades=[n]
        ctx.user_data["datos"]["recorrido"]["novedades"]=novedades
        ctx.user_data["datos"]["recorrido"]["fotos_total"]=len(media)
        for nov in novedades:
            m=nov["motivo"]
            if m!=SIN_NOV_MOTIVO:
                ctx.user_data["datos"]["mpriu"]["novedades_check"][m]={"check":True,"cantidad":ctx.user_data["datos"]["mpriu"]["novedades_check"].get(m,{}).get("cantidad",0)+1}
        await update.message.reply_text("✅ "+str(len(novedades))+" novedad(es) detectadas por la IA!")
        return await tab_menu(update, ctx)
    return TAB_NOVEDADES_IA

# ── MANGAS ────────────────────────────────────────────────────────────────────
async def recv_manga_nombre(update, ctx):
    txt=update.message.text.strip().upper()
    if txt=="FIN MANGAS":
        total=len(ctx.user_data["datos"]["mangas"])
        await update.message.reply_text("✅ "+str(total)+" manga(s) guardada(s).")
        return await tab_menu(update, ctx)
    ctx.user_data["manga_temp"]={"nombre":txt}
    teclado=InlineKeyboardMarkup([[InlineKeyboardButton("SI - Con derivacion",callback_data="manga_der_si"),InlineKeyboardButton("NO - Sin derivacion",callback_data="manga_der_no")]])
    await update.message.reply_text("✅ Nombre: "+txt+chr(10)+chr(10)+"Tiene derivacion esta manga?",reply_markup=teclado)
    return MANGA_COORDS

async def recv_manga_coords(update, ctx):
    ctx.user_data["manga_temp"]["coordenadas"]=update.message.text.strip()
    await update.message.reply_text("✅ Coordenadas: "+update.message.text.strip()+chr(10)+chr(10)+"Observacion de la manga:"+chr(10)+"Si no hay escribe: NINGUNA")
    return MANGA_OBS

async def recv_manga_obs(update, ctx):
    obs=update.message.text.strip()
    if obs.upper()=="NINGUNA": obs=""
    manga=ctx.user_data.pop("manga_temp",{})
    manga["observacion"]=obs; manga.setdefault("derivacion","NO")
    ctx.user_data["datos"]["mangas"].append(manga)
    total=len(ctx.user_data["datos"]["mangas"])
    teclado=InlineKeyboardMarkup([[InlineKeyboardButton("Terminar mangas",callback_data="tab_menu")]])
    await update.message.reply_text(
        "✅ Manga #"+str(total)+" guardada!"+chr(10)+chr(10)+
        "Nombre: "+manga.get("nombre","")+chr(10)+"Derivacion: "+manga.get("derivacion","NO")+chr(10)+
        "Coordenadas: "+manga.get("coordenadas","")+chr(10)+"Observacion: "+(obs or "NINGUNA")+chr(10)+chr(10)+
        "Nombre de la manga #"+str(total+1)+":"+chr(10)+"O escribe: FIN MANGAS",reply_markup=teclado)
    ctx.user_data["manga_temp"]={}
    return MANGA_NOMBRE

# ── HILOS ─────────────────────────────────────────────────────────────────────
async def recv_hilo_odf(update, ctx):
    ctx.user_data["datos"]["hilos"]["posicion_odf"]=update.message.text.upper()
    await update.message.reply_text("Ingresa hilos:"+chr(10)+"HILO, DESCRIPCION, ESTADO"+chr(10)+"Ejemplo: 1, TELCONET, OCUPADO"+chr(10)+"Cuando termines: FIN HILOS")
    return HILO_DATOS

async def recv_hilo_datos(update, ctx):
    if update.message.text.upper()=="FIN HILOS":
        await update.message.reply_text("✅ Hilos guardados!")
        return await tab_menu(update, ctx)
    partes=update.message.text.split(",")
    if len(partes)>=3:
        ctx.user_data["datos"]["hilos"]["filas"].append({"hilo_par":partes[0].strip(),"descripcion":partes[1].strip(),"estado":partes[2].strip().upper()})
    await update.message.reply_text("✅ Guardado. Siguiente o FIN HILOS:")
    return HILO_DATOS

# ── GENERAR EXCEL ─────────────────────────────────────────────────────────────
async def enviar_excel(update, ctx):
    try:
        msg=update.callback_query.message if update.callback_query else update.message
        if update.callback_query: await update.callback_query.answer()
        await msg.reply_text("Generando informe FOR FO 02...")
        if "datos" not in ctx.user_data: ctx.user_data["datos"]=datos_vacios()
        datos=ctx.user_data["datos"]
        if not datos["recorrido"].get("novedades"):
            nov=novedad_vacia(1); nov["motivo"]=SIN_NOV_MOTIVO; nov["remedio"]=SIN_NOV_REMEDIO
            datos["recorrido"]["novedades"]=[nov]
        xl=generar_excel(datos); nombre=nombre_archivo(datos)
        caption="FOR FO 02 generado"+chr(10)+"Ruta: "+(datos["recorrido"]["nombre_ruta"] or "SIN NOMBRE")+chr(10)+"Novedades: "+str(len(datos["recorrido"]["novedades"]))
        await msg.reply_document(document=xl,filename=nombre,caption=caption)
    except Exception as e:
        logger.error("Error generando Excel: "+str(e))
        try: await msg.reply_text("Error: "+str(e))
        except: pass
    teclado=[["Generar Informe","Nueva Ruta Base"],["Mis Rutas","Ayuda"]]
    try: await msg.reply_text("Que deseas hacer?",reply_markup=ReplyKeyboardMarkup(teclado,resize_keyboard=True))
    except: pass
    return MENU_PRINCIPAL

# ── NUEVA RUTA BASE ───────────────────────────────────────────────────────────
async def nueva_ruta(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in USUARIOS_AUTENTICADOS:
        return await start(update, ctx)
    await update.message.reply_text("Nueva Ruta Base"+chr(10)+chr(10)+"Nombre de la ruta:"+chr(10)+"Ejemplo: GOSSEAL-MACHACHI",reply_markup=ReplyKeyboardRemove())
    return NUEVA_RUTA_NOMBRE

async def recv_nueva_ruta_nombre(update, ctx):
    ctx.user_data["nueva_ruta_nombre"]=update.message.text.upper()
    nombre=ctx.user_data["nueva_ruta_nombre"]
    teclado=InlineKeyboardMarkup([[InlineKeyboardButton("Tengo el link de Mapillary",callback_data="vb_link")],[InlineKeyboardButton("Subir video directo aqui",callback_data="vb_video")],[InlineKeyboardButton("Cancelar",callback_data="tab_menu")]])
    await update.message.reply_text("Nueva Ruta Base: "+nombre+chr(10)+chr(10)+"Como quieres registrar el video base?",reply_markup=teclado)
    return NUEVA_RUTA_VIDEO

async def recv_nueva_ruta_video(update, ctx):
    nombre=ctx.user_data.get("nueva_ruta_nombre","SIN NOMBRE")
    if update.message.text and update.message.text.strip().startswith("http"):
        link=update.message.text.strip()
        RUTAS_GUARDADAS[nombre]={"nombre":nombre,"mapillary_link":link,"tipo":"mapillary","fecha":datetime.now().strftime("%d/%m/%Y %H:%M")}
        await update.message.reply_text("✅ Ruta base guardada!"+chr(10)+"Nombre: "+nombre+chr(10)+"Link: "+link)
    elif update.message.video or update.message.document:
        RUTAS_GUARDADAS[nombre]={"nombre":nombre,"tipo":"video_telegram","fecha":datetime.now().strftime("%d/%m/%Y %H:%M")}
        await update.message.reply_text("✅ Video recibido!"+chr(10)+"Ruta base guardada: "+nombre)
    else:
        await update.message.reply_text("Envia el link de Mapillary o el video."+chr(10)+"Ejemplo: https://www.mapillary.com/app/?pKey=xxx")
        return NUEVA_RUTA_VIDEO
    teclado=[["Generar Informe","Nueva Ruta Base"],["Mis Rutas","Ayuda"]]
    await update.message.reply_text("Que deseas hacer?",reply_markup=ReplyKeyboardMarkup(teclado,resize_keyboard=True))
    return MENU_PRINCIPAL

async def vb_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query=update.callback_query; await query.answer(); data=query.data
    if data=="vb_link":
        ctx.user_data["modo_video_base"]="link"
        await query.edit_message_text("Pega el link de Mapillary de tu video base:"+chr(10)+chr(10)+"Ejemplo:"+chr(10)+"https://www.mapillary.com/app/?pKey=ABC123")
        return NUEVA_RUTA_VIDEO
    elif data=="vb_video":
        ctx.user_data["modo_video_base"]="video"
        await query.edit_message_text("Envia el video de la ruta directamente."+chr(10)+"El bot lo subira a Mapillary automaticamente."+chr(10)+"Formatos: .mp4, .mov, .insv")
        return NUEVA_RUTA_VIDEO
    return MENU_PRINCIPAL

# ── MIS RUTAS / AYUDA ─────────────────────────────────────────────────────────
async def mis_rutas(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in USUARIOS_AUTENTICADOS:
        return await start(update, ctx)
    if not RUTAS_GUARDADAS:
        await update.message.reply_text("No tienes rutas base guardadas."+chr(10)+"Usa Nueva Ruta Base para registrar.")
        return MENU_PRINCIPAL
    msg="Rutas base registradas:"+chr(10)+chr(10)
    for i,(nombre,info) in enumerate(RUTAS_GUARDADAS.items(),1):
        msg+=str(i)+". "+nombre+chr(10)+"   Fecha: "+info.get("fecha","")+chr(10)
        if info.get("mapillary_link"): msg+="   Link: "+info["mapillary_link"][:50]+"..."+chr(10)
        msg+=chr(10)
    await update.message.reply_text(msg)
    return MENU_PRINCIPAL

async def ayuda(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "RecorridosIA - Ayuda"+chr(10)+chr(10)+
        "Generar Informe — abre el menu de pestanas"+chr(10)+
        "Nueva Ruta Base — registra video base de una ruta"+chr(10)+
        "Mis Rutas — lista rutas guardadas"+chr(10)+chr(10)+
        "Variables en Render:"+chr(10)+"BOT_TOKEN / GEMINI_API_KEY"+chr(10)+"MAPILLARY_TOKEN / TOTP_SECRET / DOMINIO_EMAIL")
    return MENU_PRINCIPAL

async def cancelar(update, ctx):
    ctx.user_data.clear()
    await update.message.reply_text("Cancelado.",reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END

# ── SERVIDOR WEB ──────────────────────────────────────────────────────────────
class PingHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200); self.end_headers(); self.wfile.write(b"RecorridosIA OK")
    def log_message(self,format,*args): pass

def ping_render():
    import urllib.request
    while True:
        time.sleep(720)
        try:
            url=os.getenv("RENDER_EXTERNAL_URL","")
            if url: urllib.request.urlopen(url,timeout=10); logger.info("Ping OK - bot despierto")
        except Exception as e: logger.warning("Ping error: "+str(e))

def start_web():
    port=int(os.getenv("PORT",8080))
    server=HTTPServer(("0.0.0.0",port),PingHandler)
    logger.info("Servidor web en puerto "+str(port))
    threading.Thread(target=ping_render,daemon=True).start()
    server.serve_forever()

# ── BUILD APP ─────────────────────────────────────────────────────────────────
def build_app():
    app=Application.builder().token(BOT_TOKEN).build()
    conv=ConversationHandler(
        entry_points=[CommandHandler("start",start),CommandHandler("inspeccionar",generar_informe),MessageHandler(filters.Regex("Generar Informe"),generar_informe),MessageHandler(filters.Regex("Nueva Ruta Base"),nueva_ruta),MessageHandler(filters.Regex("Mis Rutas"),mis_rutas),MessageHandler(filters.Regex("Ayuda"),ayuda)],
        states={
            ESPERANDO_TOTP:   [MessageHandler(filters.TEXT&~filters.COMMAND,handler_totp)],
            MENU_PRINCIPAL:   [MessageHandler(filters.Regex("Generar Informe"),generar_informe),MessageHandler(filters.Regex("Nueva Ruta Base"),nueva_ruta),MessageHandler(filters.Regex("Mis Rutas"),mis_rutas),MessageHandler(filters.Regex("Ayuda"),ayuda),MessageHandler(filters.TEXT&~filters.COMMAND,menu_principal)],
            TAB_MENU:         [MessageHandler(filters.TEXT&~filters.COMMAND,generar_informe)],
            TAB_CIU_HERR:     [MessageHandler(filters.TEXT&~filters.COMMAND,tab_ciu_herr)],
            TAB_CIU_EQUI:     [MessageHandler(filters.TEXT&~filters.COMMAND,tab_ciu_equi)],
            TAB_CIU_MATE:     [MessageHandler(filters.TEXT&~filters.COMMAND,tab_ciu_mate)],
            TAB_MPRIU:        [MessageHandler(filters.TEXT&~filters.COMMAND,tab_mpriu)],
            TAB_REPORTES:     [MessageHandler(filters.TEXT&~filters.COMMAND,tab_reportes)],
            TAB_NOVEDADES_IA: [MessageHandler(filters.PHOTO,tab_novedades_ia),MessageHandler(filters.TEXT&~filters.COMMAND,tab_novedades_ia)],
            MANGA_NOMBRE:     [MessageHandler(filters.TEXT&~filters.COMMAND,recv_manga_nombre)],
            MANGA_COORDS:     [MessageHandler(filters.TEXT&~filters.COMMAND,recv_manga_coords)],
            MANGA_OBS:        [MessageHandler(filters.TEXT&~filters.COMMAND,recv_manga_obs)],
            HILO_ODF:         [MessageHandler(filters.TEXT&~filters.COMMAND,recv_hilo_odf)],
            HILO_DATOS:       [MessageHandler(filters.TEXT&~filters.COMMAND,recv_hilo_datos)],
            NUEVA_RUTA_NOMBRE:[MessageHandler(filters.TEXT&~filters.COMMAND,recv_nueva_ruta_nombre)],
            NUEVA_RUTA_VIDEO: [MessageHandler(filters.TEXT|filters.VIDEO|filters.Document.ALL&~filters.COMMAND,recv_nueva_ruta_video)],
        },
        fallbacks=[CommandHandler("cancelar",cancelar)],
        allow_reentry=True,
    )
    app.add_handler(conv)
    app.add_handler(CallbackQueryHandler(tab_callback,pattern="^tab_"))
    app.add_handler(CallbackQueryHandler(tab_callback,pattern="^rep_"))
    app.add_handler(CallbackQueryHandler(vb_callback,pattern="^vb_"))
    app.add_handler(CallbackQueryHandler(tab_callback,pattern="^manga_der_"))
    return app

async def run_bot():
    app=build_app()
    await app.initialize(); await app.start(); await app.updater.start_polling()
    logger.info("RecorridosIA bot arrancando...")
    while True:
        import asyncio; await asyncio.sleep(1)

def bot_thread():
    import asyncio
    loop=asyncio.new_event_loop(); asyncio.set_event_loop(loop)
    loop.run_until_complete(run_bot())

if __name__=="__main__":
    t=threading.Thread(target=bot_thread,daemon=True); t.start()
    logger.info("RecorridosIA bot arrancando...")
    start_web()
